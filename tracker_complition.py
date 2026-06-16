#vibe

import re
from typing import Optional

from openpyxl import load_workbook


CYR_TO_LAT = str.maketrans({
    "А": "A",
    "В": "B",
    "Е": "E",
    "І": "I",
    "К": "K",
    "М": "M",
    "Н": "H",
    "О": "O",
    "Р": "P",
    "С": "C",
    "Т": "T",
    "Х": "X",
    "Ё": "E",
})


ALLOWED_PLATE_LETTERS = set("ABEKMHOPCTXI")


PLATE_RE = re.compile(
    r"(?<!\d)"
    r"(\d{4})"
    r"\s*[-–—−]?\s*"
    r"([A-ZА-ЯЁІ]{2})"
    r"\s*[-–—−]?\s*"
    r"([0-7])"
    r"(?!\d)",
    re.IGNORECASE
)


def normalize_transport_number(value: str) -> Optional[str]:
    if value is None:
        return None

    text = str(value).upper()
    text = text.replace("\u00A0", " ")

    match = PLATE_RE.search(text)
    if not match:
        return None

    digits, letters, region = match.groups()

    letters = letters.translate(CYR_TO_LAT)

    if any(letter not in ALLOWED_PLATE_LETTERS for letter in letters):
        return None

    return f"{digits} {letters}-{region}"


def find_transport_number_in_row(row_values) -> Optional[str]:
    for value in row_values:
        number = normalize_transport_number(value)
        if number:
            return number

    return None


def get_transport_numbers_from_excel(file_path: str, sheet_name: str = None) -> list[Optional[str]]:
    workbook = load_workbook(
        filename=file_path,
        read_only=True,
        data_only=True
    )

    if sheet_name:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.active

    result = [None] * (sheet.max_row + 1)

    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        result[row_index] = find_transport_number_in_row(row)

    workbook.close()

    return result


import re


CYR_TO_LAT_TRACKING = str.maketrans({
    "А": "A",
    "В": "B",
    "Е": "E",
    "К": "K",
    "М": "M",
    "Н": "H",
    "О": "O",
    "Р": "P",
    "С": "C",
    "Т": "T",
    "У": "Y",
    "Х": "X",
    "І": "I",
    "Ё": "E",
})


def get_plate_keys_for_compare(number):
    """
    Делает набор ключей для сравнения номеров.

    Нужно, чтобы такие варианты сопоставлялись:
    AB9704-7 == 9704AB7 == 9704 AB-7
    9889BA-1 == 9889 BA 1
    АВ9704-7 == AB9704-7
    """

    if not number:
        return set()

    text = str(number).upper()
    text = text.replace("\u00A0", " ")
    text = text.translate(CYR_TO_LAT_TRACKING)

    # Оставляем только буквы и цифры
    clean = re.sub(r"[^A-Z0-9]", "", text)

    if not clean:
        return set()

    keys = {clean}

    # Вид: AB97047
    match = re.fullmatch(r"([A-Z]{2})(\d{4})(\d)", clean)
    if match:
        letters, digits, region = match.groups()

        keys.add(f"{letters}{digits}{region}")      # AB97047
        keys.add(f"{digits}{letters}{region}")      # 9704AB7
        keys.add(f"BY:{digits}:{letters}:{region}") # общий ключ

    # Вид: 9704AB7
    match = re.fullmatch(r"(\d{4})([A-Z]{2})(\d)", clean)
    if match:
        digits, letters, region = match.groups()

        keys.add(f"{letters}{digits}{region}")      # AB97047
        keys.add(f"{digits}{letters}{region}")      # 9704AB7
        keys.add(f"BY:{digits}:{letters}:{region}") # общий ключ

    # Вид: E001AA7
    match = re.fullmatch(r"([A-Z])(\d{3})([A-Z]{2})(\d)", clean)
    if match:
        first_letter, digits, letters, region = match.groups()

        keys.add(f"{first_letter}{digits}{letters}{region}")
        keys.add(f"SPECIAL:{first_letter}:{digits}:{letters}:{region}")

    # Вид: AO78912
    match = re.fullmatch(r"([A-Z]{2})(\d{5})", clean)
    if match:
        letters, digits = match.groups()

        keys.add(f"{letters}{digits}")
        keys.add(f"{digits}{letters}")
        keys.add(f"SPECIAL:{letters}:{digits}")

    # Вид: 78912AO
    match = re.fullmatch(r"(\d{5})([A-Z]{2})", clean)
    if match:
        digits, letters = match.groups()

        keys.add(f"{letters}{digits}")
        keys.add(f"{digits}{letters}")
        keys.add(f"SPECIAL:{letters}:{digits}")

    # Вид: 2EHT3624
    match = re.fullmatch(r"(\d)([A-Z]{3})(\d{4})", clean)
    if match:
        first_digit, letters, digits = match.groups()

        keys.add(f"{first_digit}{letters}{digits}")
        keys.add(f"SPECIAL:{first_digit}:{letters}:{digits}")

    return keys


def build_tracking_work_by_plate(tracking_rows):
    """
    Создаёт словарь:
    ключ нормализованного номера -> последний вид работ

    Структура tracking_rows:
    [
        start_date,
        end_date,
        id_tg,
        worker_number,
        worker_name,
        city,
        type_work,
        transport_number,
        tracker_number,
        photos_count,
        comment
    ]
    """

    result = {}

    for row in tracking_rows:
        if len(row) < 8:
            continue

        type_work = row[6]
        transport_number = row[7]

        keys = get_plate_keys_for_compare(transport_number)

        for key in keys:
            # Если номер встречался несколько раз,
            # останется последний вид работ из таблицы
            result[key] = type_work

    return result


def generate_excel_tracking_compare_text(numbers_by_rows, tracking_rows):
    """
    Возвращает текст вида:

    1: AC7384-4 : Монтаж
    2: AC1284-4 : Нет информации
    3: None
    """

    tracking_work_by_plate = build_tracking_work_by_plate(tracking_rows)

    lines = []

    for row_index, found_number in enumerate(numbers_by_rows):
        # result[0] у нас пустой, чтобы индексы совпадали со строками Excel
        if row_index == 0:
            continue

        if found_number is None:
            lines.append(f"{row_index}: None")
            continue

        found_keys = get_plate_keys_for_compare(found_number)

        type_work = None

        for key in found_keys:
            if key in tracking_work_by_plate:
                type_work = tracking_work_by_plate[key]
                break

        if type_work:
            lines.append(f"<b>{row_index}: {found_number} : {type_work}</b>")
        else:
            lines.append(f"{row_index}: {found_number} : Нет инфо")

    return "\n".join(lines)









if __name__ == "__main__":
    numbers_by_rows = get_transport_numbers_from_excel("sample.xlsx")
    tracking_rows = [
    [
        "2026-06-16 11:32:05",
        "2026-06-16 11:32:46",
        "6150574145",
        "375336976948",
        "Евгений",
        "Брест",
        "Монтаж",
        "AC7384-4",
        "3646r9364",
        "2",
        "Мой комм"
    ],
        [
        "2026-06-16 11:44:37",
        "2026-06-16 11:44:59",
        "6150574145",
        "375336976948",
        "Евгений",
        "Гомель",
        "Демонтаж",
        "IX4633-6",
        "jhfbe7433984",
        "2",
        ""
    ]
]
    text = generate_excel_tracking_compare_text(numbers_by_rows, tracking_rows)
    print(text) 




